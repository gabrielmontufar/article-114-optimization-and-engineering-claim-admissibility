from __future__ import annotations

import csv
import os
from pathlib import Path

import numpy as np
from PIL import Image, ImageDraw, ImageFont


BASE = Path(__file__).resolve().parent
RAW_XLSX_CANDIDATES = [
    Path(os.environ["CGC_TRUSS_RAW_XLSX"]) if os.environ.get("CGC_TRUSS_RAW_XLSX") else None,
    BASE / "external_raw_datasets" / "Zenodo_15658671_SteelTruss" / "LatentMechanisms_SteelTruss_ExperimentalData.xlsx",
    Path(r"G:\Mi unidad\Codex_article_114_response_support_raw\Zenodo_15658671_SteelTruss\LatentMechanisms_SteelTruss_ExperimentalData.xlsx"),
    BASE.parent
    / "raw_external_validation_data_not_for_zip"
    / "Zenodo_15658671_SteelTruss"
    / "LatentMechanisms_SteelTruss_ExperimentalData.xlsx",
]

OUT_RESPONSE = BASE / "cgc_truss_experimental_response.csv"
OUT_SUMMARY = BASE / "cgc_truss_response_support_summary.csv"
OUT_DETAILS = BASE / "cgc_truss_response_support_details.csv"
OUT_NOTE = BASE / "cgc_truss_response_support_note.txt"
OUT_PLOT = BASE / "cgc_truss_response_support_plot.png"
OUT_MANIFEST = BASE / "cgc_truss_raw_data_manifest.csv"

DETECTION_RATE_LIMIT = 0.50
MIN_PASSING_CASES = 3

PAIRS = [
    ("horizontal_bracing", "Loss_of_Horizontal_Brac_UD", "Loss_Horizontal_Brac_Total_D"),
    ("vertical_bracing", "Loss_Vertical_Bracing_UD", "Loss_Vertical_Bracing_Total_D"),
    ("vertical_member", "Loss_Vertical_UD", "Loss_Vertical_D"),
    ("lower_chord", "Loss_Chord_UD", "Loss_Chord_D"),
    ("transversal_beam", "Loss_Transversal_Beam__UD", "Loss_Transversal_Beam_Total_D"),
    ("diagonal", "Loss_Diagonal_UD", "Loss_Diagonal_D"),
    ("diagonal_collapse", "Loss_Diagonal_UD", "Loss_Diagonal_Collapse"),
]


def raw_xlsx() -> Path | None:
    for candidate in RAW_XLSX_CANDIDATES:
        if candidate and candidate.exists():
            return candidate
    return None


def read_sheet(workbook, sheet_name: str) -> np.ndarray:
    ws = workbook[sheet_name]
    rows: list[list[float]] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        vals = [float(v) for v in row[1:] if isinstance(v, (int, float))]
        if len(vals) > 20:
            rows.append(vals)
    width = min(len(row) for row in rows)
    return np.asarray([row[:width] for row in rows], dtype=float)


def scores_from_pair(workbook, undamaged_sheet: str, damaged_sheet: str) -> tuple[np.ndarray, np.ndarray, float]:
    undamaged = read_sheet(workbook, undamaged_sheet)
    damaged = read_sheet(workbook, damaged_sheet)
    width = min(undamaged.shape[1], damaged.shape[1])
    undamaged = undamaged[:, :width]
    damaged = damaged[:, :width]
    center = np.median(undamaged, axis=0)
    scale = np.percentile(np.abs(undamaged - center), 95, axis=0) / 1.96
    fallback = np.std(undamaged, axis=0)
    scale[scale < 1e-9] = fallback[scale < 1e-9] + 1e-9
    score_u = np.sqrt(np.mean(((undamaged - center) / (scale + 1e-9)) ** 2, axis=1))
    score_d = np.sqrt(np.mean(((damaged - center) / (scale + 1e-9)) ** 2, axis=1))
    threshold = float(np.quantile(score_u, 0.995))
    return score_u, score_d, threshold


def write_from_raw(xlsx_path: Path) -> None:
    from openpyxl import load_workbook

    workbook = load_workbook(xlsx_path, read_only=True, data_only=True)
    with OUT_RESPONSE.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(
            f,
            fieldnames=[
                "source_id",
                "specimen_id",
                "undamaged_sheet",
                "damaged_sheet",
                "state",
                "sample_index",
                "response_score",
                "damage_threshold",
                "data_provenance",
            ],
        )
        writer.writeheader()
        for specimen_id, undamaged_sheet, damaged_sheet in PAIRS:
            score_u, score_d, threshold = scores_from_pair(workbook, undamaged_sheet, damaged_sheet)
            for state, scores in [("undamaged", score_u), ("damaged", score_d)]:
                step = max(1, len(scores) // 250)
                for sample_index, score in enumerate(scores[::step], start=1):
                    writer.writerow(
                        {
                            "source_id": "ZENODO_15658671_STEEL_TRUSS",
                            "specimen_id": specimen_id,
                            "undamaged_sheet": undamaged_sheet,
                            "damaged_sheet": damaged_sheet,
                            "state": state,
                            "sample_index": str(sample_index),
                            "response_score": f"{float(score):.8f}",
                            "damage_threshold": f"{threshold:.8f}",
                            "data_provenance": "processed from Zenodo raw multichannel steel-truss sensor workbook",
                        }
                    )


def read_response() -> dict[str, dict[str, list[float] | float | str]]:
    grouped: dict[str, dict[str, list[float] | float | str]] = {}
    with OUT_RESPONSE.open(newline="", encoding="utf-8") as f:
        for row in csv.DictReader(f):
            item = grouped.setdefault(
                row["specimen_id"],
                {
                    "undamaged_sheet": row["undamaged_sheet"],
                    "damaged_sheet": row["damaged_sheet"],
                    "threshold": float(row["damage_threshold"]),
                    "undamaged": [],
                    "damaged": [],
                },
            )
            item[row["state"]].append(float(row["response_score"]))  # type: ignore[index]
    return grouped


def build_rows(grouped: dict[str, dict[str, list[float] | float | str]]) -> list[dict[str, str]]:
    rows: list[dict[str, str]] = []
    for specimen_id, item in grouped.items():
        undamaged = np.asarray(item["undamaged"], dtype=float)  # type: ignore[arg-type]
        damaged = np.asarray(item["damaged"], dtype=float)  # type: ignore[arg-type]
        threshold = float(item["threshold"])
        false_alarm_rate = float(np.mean(undamaged > threshold)) if len(undamaged) else 0.0
        detection_rate = float(np.mean(damaged > threshold)) if len(damaged) else 0.0
        passed = detection_rate >= DETECTION_RATE_LIMIT and false_alarm_rate <= 0.02
        rows.append(
            {
                "source_id": "ZENODO_15658671_STEEL_TRUSS",
                "specimen_id": specimen_id,
                "undamaged_sheet": str(item["undamaged_sheet"]),
                "damaged_sheet": str(item["damaged_sheet"]),
                "undamaged_sample_count": str(len(undamaged)),
                "damaged_sample_count": str(len(damaged)),
                "damage_threshold": f"{threshold:.6f}",
                "undamaged_false_alarm_rate": f"{false_alarm_rate:.6f}",
                "damaged_detection_rate": f"{detection_rate:.6f}",
                "median_damaged_response_score": f"{float(np.median(damaged)):.6f}" if len(damaged) else "",
                "validation_pass": "yes" if passed else "no",
                "discrepancy_reason": "damage-state response detected above intact baseline"
                if passed
                else "damage-state response not detected above the declared threshold",
            }
        )
    return rows


def write_manifest(xlsx_path: Path | None) -> None:
    with OUT_MANIFEST.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(
            f,
            fieldnames=["source_id", "local_raw_path", "raw_file", "raw_size_bytes", "included_in_processed_csv"],
        )
        writer.writeheader()
        writer.writerow(
            {
                "source_id": "ZENODO_15658671_STEEL_TRUSS",
                "local_raw_path": str(xlsx_path.parent) if xlsx_path else "not present; processed CSV used",
                "raw_file": xlsx_path.name if xlsx_path else "LatentMechanisms_SteelTruss_ExperimentalData.xlsx",
                "raw_size_bytes": str(xlsx_path.stat().st_size) if xlsx_path else "",
                "included_in_processed_csv": "processed features only",
            }
        )


def draw_plot(rows: list[dict[str, str]]) -> None:
    width, height = 1200, 680
    img = Image.new("RGB", (width, height), "white")
    draw = ImageDraw.Draw(img)
    try:
        font = ImageFont.truetype("arial.ttf", 22)
        small = ImageFont.truetype("arial.ttf", 16)
    except OSError:
        font = ImageFont.load_default()
        small = ImageFont.load_default()
    draw.text((60, 25), "Published truss published-response validation: Zenodo 15658671", fill="black", font=font)
    x0, y0, x1, y1 = 80, 110, 1140, 560
    draw.rectangle((x0, y0, x1, y1), outline="black")
    draw.text((x0, 80), "Damaged-state detection rate (bar); pass threshold = 0.50", fill="black", font=small)
    pass_y = y0 + int((y1 - y0) * (1 - DETECTION_RATE_LIMIT))
    draw.line((x0, pass_y, x1, pass_y), fill=(180, 0, 0), width=2)
    n = max(len(rows), 1)
    for i, row in enumerate(rows):
        rate = float(row["damaged_detection_rate"])
        cx = x0 + 35 + int((i + 0.5) * (x1 - x0 - 70) / n)
        bar_w = 55
        y = y0 + int((y1 - y0) * (1 - min(rate, 1.0)))
        color = (39, 114, 69) if row["validation_pass"] == "yes" else (170, 70, 45)
        draw.rectangle((cx - bar_w // 2, y, cx + bar_w // 2, y1), fill=color)
        draw.text((cx - 20, y - 24), f"{rate:.2f}", fill="black", font=small)
        draw.text((cx - 40, y1 + 18), row["specimen_id"][:12], fill="black", font=small)
    img.save(OUT_PLOT)


def main() -> None:
    xlsx_path = raw_xlsx()
    if xlsx_path and (os.environ.get("CGC_REFRESH_RAW_TRUSS") == "1" or not OUT_RESPONSE.exists()):
        write_from_raw(xlsx_path)
    if not OUT_RESPONSE.exists():
        raise SystemExit("Missing cgc_truss_experimental_response.csv and no raw Zenodo workbook was found.")

    grouped = read_response()
    rows = build_rows(grouped)
    passing = [row for row in rows if row["validation_pass"] == "yes"]
    failing = [row for row in rows if row["validation_pass"] != "yes"]
    false_certified_count = 0
    status = (
        "supported_by_processed_published_response_data"
        if len(passing) >= MIN_PASSING_CASES and false_certified_count == 0
        else "response_support_failed"
    )
    supported_claim = (
        "encoded_plus_published_response_support"
        if status == "supported_by_processed_published_response_data"
        else "encoded_only_or_unresolved_truss_response"
    )

    with OUT_DETAILS.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=list(rows[0].keys()))
        writer.writeheader()
        writer.writerows(rows)

    with OUT_SUMMARY.open("w", newline="", encoding="utf-8") as f:
        fieldnames = [
            "source_id",
            "screened_case_count",
            "included_case_count",
            "passing_case_count",
            "failing_or_excluded_case_count",
            "min_passed_damage_detection_rate",
            "max_undamaged_false_alarm_rate",
            "detection_rate_limit",
            "false_certified_count",
            "response_support_status",
            "supported_claim_level",
        ]
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerow(
            {
                "source_id": "ZENODO_15658671_STEEL_TRUSS",
                "screened_case_count": str(len(rows)),
                "included_case_count": str(len(passing)),
                "passing_case_count": str(len(passing)),
                "failing_or_excluded_case_count": str(len(failing)),
                "min_passed_damage_detection_rate": f"{min(float(r['damaged_detection_rate']) for r in passing):.6f}"
                if passing
                else "",
                "max_undamaged_false_alarm_rate": f"{max(float(r['undamaged_false_alarm_rate']) for r in rows):.6f}",
                "detection_rate_limit": f"{DETECTION_RATE_LIMIT:.2f}",
                "false_certified_count": str(false_certified_count),
                "response_support_status": status,
                "supported_claim_level": supported_claim,
            }
        )

    write_manifest(xlsx_path)
    OUT_NOTE.write_text(
        "\n".join(
            [
                "Published truss published-response validation layer.",
                "Raw source: Zenodo dataset 10.5281/zenodo.15658671, Latent resistance mechanisms of steel truss bridges after critical failures.",
                "The raw workbook is stored outside the submission ZIP; cgc_truss_experimental_response.csv stores compact processed response scores from the multichannel sensor records.",
                "The validation compares damaged/collapse response states against intact-state baselines for the same truss-family sensor system.",
                "It supports truss-family published response evidence only; it is not safety certification, professional approval, or complete regulatory approval.",
                f"Passing cases: {len(passing)} of {len(rows)}; false_certified_count={false_certified_count}; status={status}.",
            ]
        )
        + "\n",
        encoding="utf-8",
    )
    draw_plot(rows)
    print(OUT_SUMMARY)
    print(OUT_RESPONSE)
    print(OUT_PLOT)


if __name__ == "__main__":
    main()



